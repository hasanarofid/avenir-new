import pandas as pd
import json
import sys
import math
import re
import openpyxl

def get_entity_key(kind, ticker_or_name):
    if kind == 'issuer':
        return f"E:{ticker_or_name}"
    else:
        return f"I:{str(ticker_or_name).strip().upper()}"

def num(v):
    if v == '' or v is None: return None
    try:
        s = str(v).replace(',', '').strip()
        n = float(s)
        return n if not math.isnan(n) else None
    except ValueError:
        return None

def parse_ksei_workbook(file_path):
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    row_iter = ws.iter_rows(values_only=True)
    
    hIdx = -1
    hdr = []
    
    for i in range(12):
        try:
            r = next(row_iter)
            r = ["" if x is None else str(x).strip() for x in r]
            if any(c == 'Kode Efek' for c in r):
                hIdx = i
                hdr = r
                break
        except StopIteration:
            break
            
    if hIdx < 0:
        wb.close()
        raise ValueError('Format tidak dikenali: baris header "Kode Efek" tak ditemukan.')
    
    datePrev = ''
    dateNow = ''
    dateCols = []
    for idx, c in enumerate(hdr):
        if 'Kepemilikan Per' in c:
            dateCols.append({'idx': idx, 'label': c.replace('Kepemilikan Per', '').strip()})
    
    if len(dateCols) >= 2:
        datePrev = dateCols[0]['label']
        dateNow = dateCols[1]['label']
    elif len(dateCols) == 1:
        dateNow = dateCols[0]['label']
        
    issuers = {}
    cur = None
    
    def flush():
        nonlocal cur
        if not cur: return
        totchg = 0
        moved = []
        for x in cur['rows']:
            if x['chg']:
                totchg += x['chg']
                if abs(x['chg']) > 0:
                    moved.append(x)
        
        broker = cur['broker']
        if moved:
            moved.sort(key=lambda a: abs(a['chg']), reverse=True)
            broker = moved[0]['broker'] or broker
            
        sharesPrev = cur['sharesPrev']
        if sharesPrev is None and cur['sharesNow'] is not None:
            sharesPrev = cur['sharesNow'] - totchg
            
        if cur['ticker'] not in issuers:
            issuers[cur['ticker']] = []
            
        issuers[cur['ticker']].append({
            'investor': cur['investor'],
            'broker': broker,
            'change': totchg,
            'pctNow': cur['pctNow'],
            'pctPrev': cur['pctPrev'],
            'sharesNow': cur['sharesNow'],
            'sharesPrev': sharesPrev,
            'type': cur['type'],
            'domicile': cur['domicile'],
            'brokerCount': len(cur['rows'])
        })

    try:
        next(row_iter)
    except StopIteration:
        pass

    for r_raw in row_iter:
        r = ["" if x is None else x for x in r_raw]
        no = r[0] if len(r) > 0 else None
        ticker = r[1] if len(r) > 1 else None
        if no == '' or ticker == '': continue
        
        t = str(ticker).strip()
        investor = str(r[4]).strip() if len(r) > 4 else ''
        broker = str(r[3]).strip() if len(r) > 3 else ''
        chg = num(r[17]) if len(r) > 17 else None
        
        if investor:
            flush()
            cur = {
                'ticker': t,
                'investor': investor,
                'broker': broker,
                'pctNow': num(r[16]) if len(r) > 16 else None,
                'pctPrev': num(r[13]) if len(r) > 13 else None,
                'sharesNow': num(r[15]) if len(r) > 15 else None,
                'sharesPrev': num(r[12]) if len(r) > 12 else None,
                'type': str(r[10]).strip() if len(r) > 10 else '',
                'domicile': str(r[9]).strip() if len(r) > 9 else '',
                'rows': [{'broker': broker, 'chg': chg}]
            }
        elif cur and cur['ticker'] == t:
            cur['rows'].append({'broker': broker, 'chg': chg})
            
    flush()
    wb.close()
    return {'dateNow': dateNow, 'datePrev': datePrev, 'issuers': issuers, 'source': 'KSEI >=5%'}

def parse_monthly_workbook(file_path):
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    row_iter = ws.iter_rows(values_only=True)
    
    buffer = []
    for i in range(12):
        try:
            r = next(row_iter)
            buffer.append(["" if x is None else x for x in r])
        except StopIteration:
            break
            
    flat = "||".join(["|".join([str(c) for c in r]) for r in buffer[:8]]).upper()
    
    def get_date_from_buffer(startIdx, col):
        for i in range(startIdx, min(len(buffer), startIdx + 6)):
            v = buffer[i][col] if col < len(buffer[i]) else ''
            if v:
                m = re.search(r'(\d{4})-(\d{2})-(\d{2})', str(v))
                if m: return m.group(0)
                try:
                    d = pd.to_datetime(v)
                    return d.strftime('%Y-%m-%d')
                except:
                    pass
        return None

    def find_header_in_buffer(token):
        for i in range(len(buffer)):
            if any(str(c).strip().upper() == token for c in buffer[i]):
                return i
        return -1

    if 'INVESTOR_NAME' in flat or 'INVESTOR NAME' in flat:
        h = find_header_in_buffer('SHARE_CODE')
        if h < 0: h = find_header_in_buffer('SHARE CODE')
        hi = h if h >= 0 else 5
        date = get_date_from_buffer(hi+1, 0)
        
        issuers = {}
        for i in range(hi+1, len(buffer)):
            r = buffer[i]
            if len(r) < 2 or r[1] == '': continue
            t = str(r[1]).strip()
            if t not in issuers: issuers[t] = []
            issuers[t].append({
                'investor': str(r[3]).strip() if len(r) > 3 else '',
                'cls': str(r[4]).strip() if len(r) > 4 else '',
                'lf': str(r[5]).strip() if len(r) > 5 else '',
                'domicile': str(r[7]).strip() if len(r) > 7 else '',
                'shares': num(r[10]) if len(r) > 10 else None,
                'pct': num(r[11]) if len(r) > 11 else None
            })
        
        for r_raw in row_iter:
            r = ["" if x is None else x for x in r_raw]
            if len(r) < 2 or r[1] == '': continue
            t = str(r[1]).strip()
            if t not in issuers: issuers[t] = []
            issuers[t].append({
                'investor': str(r[3]).strip() if len(r) > 3 else '',
                'cls': str(r[4]).strip() if len(r) > 4 else '',
                'lf': str(r[5]).strip() if len(r) > 5 else '',
                'domicile': str(r[7]).strip() if len(r) > 7 else '',
                'shares': num(r[10]) if len(r) > 10 else None,
                'pct': num(r[11]) if len(r) > 11 else None
            })
        wb.close()
        return {'kind': 'satu', 'date': date, 'issuers': issuers}
        
    if 'TOTAL SCRIPLESS' in flat and 'GOVERNMENT' in flat and 'INDIVIDUAL' in flat:
        hi = find_header_in_buffer('SHARE CODE')
        if hi < 0: hi = find_header_in_buffer('SHARE_CODE')
        hr = buffer[hi] if hi >= 0 else []
        cats = []
        totalCol = -1
        for c in range(3, len(hr)):
            nm = str(hr[c]).strip()
            if nm == 'TOTAL SCRIPLESS':
                totalCol = c
            elif nm:
                cats.append((c, nm))
                
        date = get_date_from_buffer(hi+1, 0)
        issuers = {}
        
        def process_klas_row(r):
            if len(r) < 2 or r[1] == '': return
            comp = {}
            for c, nm in cats:
                v = num(r[c]) if c < len(r) else None
                if v and v > 0: comp[nm] = v
            t = str(r[1]).strip()
            issuers[t] = {
                'total': num(r[totalCol]) if totalCol >= 0 and totalCol < len(r) else None,
                'comp': comp
            }
            
        for i in range(hi+1, len(buffer)):
            process_klas_row(buffer[i])
            
        for r_raw in row_iter:
            r = ["" if x is None else x for x in r_raw]
            process_klas_row(r)
            
        wb.close()
        return {'kind': 'klas', 'date': date, 'issuers': issuers}
        
    if 'DOMESTIC' in flat and 'FOREIGN' in flat:
        hi = find_header_in_buffer('STOCK_CODE')
        if hi < 0: hi = 1
        hr = buffer[hi] if hi >= 0 else []
        totalCol = next((i for i, c in enumerate(hr) if str(c).strip().upper() == 'TOTAL SCRIPLESS'), -1)
        foreignStart = next((i for i, c in enumerate(hr) if str(c).strip().upper() == 'FOREIGN'), -1)
        
        while len(buffer) < hi + 3 + 6:
            try:
                r = next(row_iter)
                buffer.append(["" if x is None else x for x in r])
            except StopIteration:
                break
                
        date = get_date_from_buffer(hi+3, 0)
        issuers = {}
        
        def process_tipe_row(r):
            if len(r) < 2 or r[1] == '': return
            dom = 0
            foreign = 0
            tot = num(r[totalCol]) if totalCol >= 0 and totalCol < len(r) else 0
            end_col = totalCol if totalCol > 0 else len(r)
            for c in range(3, end_col):
                v = num(r[c]) if c < len(r) else 0
                if not v: v = 0
                if foreignStart > 0 and c >= foreignStart:
                    foreign += v
                else:
                    dom += v
            t = str(r[1]).strip()
            issuers[t] = {'domestic': dom, 'foreign': foreign, 'total': tot}
            
        for i in range(hi+3, len(buffer)):
            process_tipe_row(buffer[i])
            
        for r_raw in row_iter:
            r = ["" if x is None else x for x in r_raw]
            process_tipe_row(r)
            
        wb.close()
        return {'kind': 'tipe', 'date': date, 'issuers': issuers}
        
    wb.close()
    raise ValueError('Jenis file bulanan tidak dikenali (bukan satu-persen/klasifikasi/tipe).')

print("Syntax OK")
