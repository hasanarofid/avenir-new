import pandas as pd
from openpyxl import Workbook

file_path = '/home/hasanarofid/Documents/hasanarofid/avenir/kebutuhan/Data Foreign Flow.xlsx'
wb = Workbook()
ws = wb.active
ws.title = "Data"

ws['B2'] = 'Tanggal'
ws['C2'] = 'Foreign Net Flow'
ws['D2'] = 'Market Value'
ws['E2'] = 'Price'
ws['F2'] = 'Open'
ws['G2'] = 'High'
ws['H2'] = 'Low'
ws['I2'] = 'Volume'
ws['J2'] = 'Change %'

ws['B3'] = '10/07/2026'
# Use dots for decimals!
ws['C3'] = '-284.40B'
ws['D3'] = '8.86T'
ws['E3'] = 5924
ws['F3'] = 5924
ws['G3'] = 5924
ws['H3'] = 5924
ws['I3'] = '0B'
ws['J3'] = '0.00%'

wb.save(file_path)
print("File berhasil diperbaiki dengan format desimal titik.")
