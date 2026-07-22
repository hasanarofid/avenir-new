import pandas as pd
from openpyxl import load_workbook

file_path = '/home/hasanarofid/Documents/hasanarofid/avenir/kebutuhan/Data Foreign Flow.xlsx'

# Load the workbook
wb = load_workbook(file_path)
ws = wb['Data']

# Find the next empty row
next_row = ws.max_row + 1

# Columns based on the image:
# B = Tanggal, C = Foreign Net Flow, D = Market Value, E = Price, F = Open, G = High, H = Low, I = Volume, J = Change %
ws[f'B{next_row}'] = '10/07/2026'
ws[f'C{next_row}'] = '-284,40B'
ws[f'D{next_row}'] = '8,86T'
ws[f'E{next_row}'] = 5924
ws[f'F{next_row}'] = 5924
ws[f'G{next_row}'] = 5924
ws[f'H{next_row}'] = 5924
ws[f'I{next_row}'] = '0B'
ws[f'J{next_row}'] = '0.00%'

wb.save(file_path)
print("Berhasil ditambahkan!")
