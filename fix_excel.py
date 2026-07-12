import pandas as pd
from openpyxl import load_workbook

file_path = '/home/hasanarofid/Documents/hasanarofid/avenir/kebutuhan/Data Foreign Flow.xlsx'
wb = load_workbook(file_path)
ws = wb['Data']

# Find the real max row by checking column B (Tanggal) starting from bottom
real_max_row = 1
for row in range(ws.max_row, 1, -1):
    if ws[f'B{row}'].value is not None:
        real_max_row = row
        break

# The row with 10/07/2026 was added at the very bottom (like 999).
# We'll just delete everything after the real_max_row (which might be the one we just added if it wasn't empty)
# Let's be safer. First, find where 09/07/2026 is.
last_valid_row = 1
for row in range(2, ws.max_row + 1):
    val = ws[f'B{row}'].value
    if val and '09/07/2026' in str(val):
        last_valid_row = row
        break

print(f"Row for 09/07/2026 is {last_valid_row}")

next_row = last_valid_row + 1
ws[f'B{next_row}'] = '10/07/2026'
ws[f'C{next_row}'] = '-284,40B'
ws[f'D{next_row}'] = '8,86T'
ws[f'E{next_row}'] = 5924
ws[f'F{next_row}'] = 5924
ws[f'G{next_row}'] = 5924
ws[f'H{next_row}'] = 5924
ws[f'I{next_row}'] = '0B'
ws[f'J{next_row}'] = '0.00%'

# Delete rows after next_row
for row in range(ws.max_row, next_row, -1):
    ws.delete_rows(row)

wb.save(file_path)
print(f"Fixed! Data inserted at row {next_row} and extra rows deleted.")
